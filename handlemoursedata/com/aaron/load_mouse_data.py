# -*- coding: utf-8 -*-
'''
Created on 2019-04-19
@FileName: load_mouse_data.py
@Description: 处理行业参数数据
@author: 'Aaron.Qiu'
@version V1.0.0
'''

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

from com.aaron.bean.component_property import ComponentProperty
from com.aaron.bean.component_prop import ComponentProp
from com.aaron.dao.component_dao import addComponent


def load_excel():
    try:
        workbook_ = load_workbook("E:\Mouser原始数据.xlsx")
        sheetnames = workbook_.get_sheet_names()  # 获得表单名字
        # print(sheetnames)
        sheet = workbook_.get_sheet_by_name(sheetnames[0])
        for row_index in range(2,28678):
            component = ComponentProperty()
            propKey = None
            prop_list = None
            for col_index in range(1,9):
                cell_value = sheet.cell(row=row_index, column=col_index).value
                if cell_value is None:
                    cell_value = '-'
                print(cell_value,end='\t')
                if(col_index == 1):
                    component.categoryTypeNum = str(cell_value)
                if(col_index == 2):
                    component.category1 = cell_value
                if (col_index == 3):
                    component.category2 = cell_value
                if(col_index == 4):
                    component.category3 = cell_value
                if (col_index == 5):
                    component.category4 = cell_value
                if (col_index == 7):
                    propKey = cell_value
                if (col_index == 8):
                    propValue = cell_value
                    try:
                        if(isinstance(propValue,int)):
                            propValue = str(propValue)

                        if (isinstance(propValue, float)):
                            propValue = str(propValue)

                        prop_list = propValue.split('^')
                    except Exception as e:
                        print("=====解析异常=====")
                        print(type(propValue))
                        # raise e
                    # prop_str = ''.join(prop_list)
                    # print(prop_str)

            if(propKey is not None):
                cp_list = []
                if(prop_list is not None):
                    for prop in prop_list:
                        cp = ComponentProp(propKey,prop)
                        cp_list.append(cp)

                # prop_dic = {propKey:prop_list}
                component.setPropList(cp_list)
            print("")
            print(component)
            componentList = component.getPropList()
            for cp in component.getPropList():
                print(cp)

            # 数据入库
            addComponent(component)


            # print(component.getPropDic())
        # print(sheet['A2'].value)
    except Exception as e:
        print(e)
        raise e



